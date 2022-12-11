from openpyxl import Workbook, load_workbook
import os
from time import sleep

input_path = input("input path: ").replace('\\', '/') 
output_path = input("output path: ").replace('\\', '/')
file_name = input('file save Name: ').replace('.xlsx', '')
input_path =  input_path if input_path[-1] == '/' else f'{input_path}/'
output_path = output_path if input_path[-1] == '/' else f'{input_path}/'
print("out put file show in same dir")
create_sheet = Workbook()
for file in os.listdir(input_path):
    if '.xlsx' in file:
        read_file = load_workbook(input_path +  file).active
        hello =  create_sheet.create_sheet(file[:-5])
        for row in range(1, len(read_file['A'])):
            hello.append([i.value for i in read_file[row:row]])
del create_sheet['Sheet']
create_sheet.save(f'{output_path}/{file_name}.xlsx')
